from __future__ import unicode_literals
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import datetime
import random
import seaborn as sns
import sys

sys.tracebacklimit = 0

class OrgTable:
    def __init__(self, source, output):
        # Define Variables
        self.sourceFilename = source
        self.outputFilename = output
        self.toPrint = []
        self.line = 5
        self.maxlevel = 0
        # Read the Source File
        try:
            wb = xl.load_workbook(self.sourceFilename)
            wb.active = 1
            self.sheet=wb.active
            self.max_row=self.sheet.max_row
        except FileNotFoundError as e:
            print(f"!ERROR: Could not open the source file please validate the path and name {e}")
            raise
        # Load the Out file to Write
        try:
            self.BuildOutFile()
            self.outWB = xl.load_workbook(self.outputFilename)
            self.outSheet=self.outWB.active
        except FileNotFoundError as e:
            print(f"!ERROR: Could not create or load the output file please validate the path and name {e}")
            raise
    # Create the output file
    def BuildOutFile(self):
        outputFile=Workbook()
        filepath=self.outputFilename
        outputFile.save(filepath)
        return None
    # Returns the Root user data
    def getRoot(self):
        rootUID=[]
        for i in range(2,self.max_row+1):
            worksfor=self.sheet.cell(row=i,column=3).value
            if worksfor == "":
                userid=self.sheet.cell(row=i,column=1).value
                name=self.sheet.cell(row=i,column=2).value
                role=self.sheet.cell(row=i, column=4).value
                location=self.sheet.cell(row=i, column=5).value
                orgName=self.sheet.cell(row=i, column=6).value
                rootUID.append({
                    'userid' : userid,
                    'name' : name,
                    'worksfor' : worksfor,
                    'role': role,
                    'location' : location,
                    'orgName' :  orgName
                })
        return rootUID
    # Get the list of subordinates from the given uid
    def getSubs(self,boss):
        subs=[]
        for x in range(2,self.max_row+1):
            userid=self.sheet.cell(row=x,column=1).value
            name=self.sheet.cell(row=x,column=2).value
            worksfor=self.sheet.cell(row=x,column=3).value
            role=self.sheet.cell(row=x, column=4).value
            location=self.sheet.cell(row=x, column=5).value
            orgName=self.sheet.cell(row=x, column=6).value
            #print(f"{who.value} reports to {worksfor.value}")
            if worksfor == boss:
                found = 1
                # Add the user data to the list of subordinates
                subs.append({
                    'userid' : userid,
                    'name' : name,
                    'worksfor' : worksfor,
                    'role': role,
                    'location' : location,
                    'orgName' :  orgName
                })
        return subs
    # Write data to the output xlsx file    
    def writeXlsx(self,line,data):
        x=1
        for field in data:
            self.outSheet.cell(row=line, column=x).value = field
            x=x+1
        self.line = self.line+1
        return None
    def addHeaders(self):
        print(f"{datetime.datetime.now()} Adding Headers and Format to the file..")
        HeaderstoPrint=[]
        for i in range(self.maxlevel+1):
            HeaderstoPrint.append('Unique Identifier')
            HeaderstoPrint.append('Name')
            HeaderstoPrint.append('Reports To')
            HeaderstoPrint.append('Role')
            HeaderstoPrint.append('Location')
            HeaderstoPrint.append('Organization Name')
        x=1
        for field in HeaderstoPrint:
            self.outSheet.cell(row=4, column=x).value = field
            self.outSheet.cell(row=4, column=x).fill = PatternFill(start_color='3399FF', end_color='3399FF', fill_type = "solid")
            self.outSheet.cell(row=4, column=x).font = Font(bold=True,color='FFFFFF')
            self.outSheet.cell(row=4, column=x).alignment = Alignment(horizontal="center", vertical="center")
            #self.outSheet.cell(row=4, column=x).fill = PatternFill("solid", start_color="#0080FF")
            x=x+1
        start=1
        end=6
        color_list=sns.color_palette("deep").as_hex()
        for i in range(self.maxlevel+1):
            color = random.choice(color_list)[1:].upper()
            self.outSheet.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            self.outSheet.cell(row=2, column=start).fill = PatternFill(start_color=color, end_color=color, fill_type = "solid")
            self.outSheet.cell(row=2, column=start).value = f"Nivel {i}"
            self.outSheet.cell(row=2, column=start).font = Font(bold=True,color='FFFFFF')
            self.outSheet.cell(row=2, column=start).alignment = Alignment(horizontal="center", vertical="center")
            start=start+6
            end=end+6
        self.outSheet.freeze_panes = 'C2'
        self.outWB.save(self.outputFilename)
        print(f"{datetime.datetime.now()} Completed..")
        return None
    # Process the source data    
    def ProcessData(self):
        Level0 = []
        Level1 = []
        Level2 = []
        Level3 = []
        Level4 = []
        Level5 = []
        Level6 = []
        Level7 = []
        Level8 = []
        Level9 = []
        print(f"{datetime.datetime.now()} Processing the input file")
        # Get the root UID and add it to the output file
        rootUID = self.getRoot()
        self.toPrint.append(rootUID[0]['userid'])
        self.toPrint.append(rootUID[0]['name'])
        self.toPrint.append(rootUID[0]['worksfor'])
        self.toPrint.append(rootUID[0]['role'])
        self.toPrint.append(rootUID[0]['location'])
        self.toPrint.append(rootUID[0]['orgName'])
        self.writeXlsx(self.line,self.toPrint)
        self.toPrint.clear()
        # Iterate the users to find the leadership tree
        Level0 = self.getSubs(rootUID[0]['userid'])
        for uid0 in Level0:
            if not self.toPrint:
                self.toPrint.append(uid0['userid'])
                self.toPrint.append(uid0['name'])
                self.toPrint.append(uid0['worksfor'])
                self.toPrint.append(uid0['role'])
                self.toPrint.append(uid0['location'])
                self.toPrint.append(uid0['orgName'])
                self.writeXlsx(self.line,self.toPrint)
            else:
                self.toPrint.clear()
                self.toPrint.append(uid0['userid'])
                self.toPrint.append(uid0['name'])
                self.toPrint.append(uid0['worksfor'])
                self.toPrint.append(uid0['role'])
                self.toPrint.append(uid0['location'])
                self.toPrint.append(uid0['orgName'])
                self.writeXlsx(self.line,self.toPrint)
            Level1 = self.getSubs(uid0['userid'])
            for uid1 in Level1:
                if self.maxlevel<=1:
                    self.maxlevel=1
                if len(self.toPrint) == 6:
                    self.toPrint.append(uid1['userid'])
                    self.toPrint.append(uid1['name'])
                    self.toPrint.append(uid1['worksfor'])
                    self.toPrint.append(uid1['role'])
                    self.toPrint.append(uid1['location'])
                    self.toPrint.append(uid1['orgName'])
                    self.writeXlsx(self.line,self.toPrint)
                elif len(self.toPrint) > 6:
                    #toPrint.clear()
                    self.toPrint = self.toPrint[:6]
                    self.toPrint.append(uid1['userid'])
                    self.toPrint.append(uid1['name'])
                    self.toPrint.append(uid1['worksfor'])
                    self.toPrint.append(uid1['role'])
                    self.toPrint.append(uid1['location'])
                    self.toPrint.append(uid1['orgName'])
                    self.writeXlsx(self.line,self.toPrint)
                Level2 = self.getSubs(uid1['userid'])
                for uid2 in Level2:
                    if self.maxlevel<=2:
                        self.maxlevel=2
                    if len(self.toPrint) == 12:
                        self.toPrint.append(uid2['userid'])
                        self.toPrint.append(uid2['name'])
                        self.toPrint.append(uid2['worksfor'])
                        self.toPrint.append(uid2['role'])
                        self.toPrint.append(uid2['location'])
                        self.toPrint.append(uid2['orgName'])
                        #print(f"!L2: {toPrint}")
                        self.writeXlsx(self.line,self.toPrint)
                    elif len(self.toPrint) > 12:
                        self.toPrint = self.toPrint[:12]
                        self.toPrint.append(uid2['userid'])
                        self.toPrint.append(uid2['name'])
                        self.toPrint.append(uid2['worksfor'])
                        self.toPrint.append(uid2['role'])
                        self.toPrint.append(uid2['location'])
                        self.toPrint.append(uid2['orgName'])
                        self.writeXlsx(self.line,self.toPrint)        
                    Level3 = self.getSubs(uid2['userid'])
                    for uid3 in Level3:
                        if self.maxlevel<=3:
                            self.maxlevel=3
                        if len(self.toPrint) == 18:
                            self.toPrint.append(uid3['userid'])
                            self.toPrint.append(uid3['name'])
                            self.toPrint.append(uid3['worksfor'])
                            self.toPrint.append(uid3['role'])
                            self.toPrint.append(uid3['location'])
                            self.toPrint.append(uid3['orgName'])
                            self.writeXlsx(self.line,self.toPrint)            
                        elif len(self.toPrint) > 18:
                            self.toPrint = self.toPrint[:18]
                            self.toPrint.append(uid3['userid'])
                            self.toPrint.append(uid3['name'])
                            self.toPrint.append(uid3['worksfor'])
                            self.toPrint.append(uid3['role'])
                            self.toPrint.append(uid3['location'])
                            self.toPrint.append(uid3['orgName'])
                            self.writeXlsx(self.line,self.toPrint)
                        Level4 = self.getSubs(uid3['userid'])
                        for uid4 in Level4:
                            if self.maxlevel<=4:
                                self.maxlevel=4
                            if len(self.toPrint) == 24:
                                self.toPrint.append(uid4['userid'])
                                self.toPrint.append(uid4['name'])
                                self.toPrint.append(uid4['worksfor'])
                                self.toPrint.append(uid4['role'])
                                self.toPrint.append(uid4['location'])
                                self.toPrint.append(uid4['orgName'])
                                self.writeXlsx(self.line,self.toPrint)            
                            elif len(self.toPrint) > 24:
                                self.toPrint = self.toPrint[:24]
                                self.toPrint.append(uid4['userid'])
                                self.toPrint.append(uid4['name'])
                                self.toPrint.append(uid4['worksfor'])
                                self.toPrint.append(uid4['role'])
                                self.toPrint.append(uid4['location'])
                                self.toPrint.append(uid4['orgName'])
                                self.writeXlsx(self.line,self.toPrint)
                            Level5 = self.getSubs(uid4['userid'])
                            for uid5 in Level5:
                                if self.maxlevel<=5:
                                    self.maxlevel=5
                                if len(self.toPrint) == 30:
                                    self.toPrint.append(uid5['userid'])
                                    self.toPrint.append(uid5['name'])
                                    self.toPrint.append(uid5['worksfor'])
                                    self.toPrint.append(uid5['role'])
                                    self.toPrint.append(uid5['location'])
                                    self.toPrint.append(uid5['orgName'])
                                    self.writeXlsx(self.line,self.toPrint)            
                                elif len(self.toPrint) > 30:
                                    self.toPrint = self.toPrint[:30]
                                    self.toPrint.append(uid5['userid'])
                                    self.toPrint.append(uid5['name'])
                                    self.toPrint.append(uid5['worksfor'])
                                    self.toPrint.append(uid5['role'])
                                    self.toPrint.append(uid5['location'])
                                    self.toPrint.append(uid5['orgName'])
                                    self.writeXlsx(self.line,self.toPrint)
                                Level6 = self.getSubs(uid5['userid'])
                                for uid6 in Level6:
                                    if self.maxlevel<=6:
                                        self.maxlevel=6
                                    if len(self.toPrint) == 36:
                                        self.toPrint.append(uid6['userid'])
                                        self.toPrint.append(uid6['name'])
                                        self.toPrint.append(uid6['worksfor'])
                                        self.toPrint.append(uid6['role'])
                                        self.toPrint.append(uid6['location'])
                                        self.toPrint.append(uid6['orgName'])
                                        self.writeXlsx(self.line,self.toPrint)            
                                    elif len(self.toPrint) > 36:
                                        self.toPrint = self.toPrint[:36]
                                        self.toPrint.append(uid6['userid'])
                                        self.toPrint.append(uid6['name'])
                                        self.toPrint.append(uid6['worksfor'])
                                        self.toPrint.append(uid6['role'])
                                        self.toPrint.append(uid6['location'])
                                        self.toPrint.append(uid6['orgName'])
                                        self.writeXlsx(self.line,self.toPrint)
                                    Level7 = self.getSubs(uid6['userid'])
                                    for uid7 in Level7:
                                        if self.maxlevel<=7:
                                            self.maxlevel=7
                                        if len(self.toPrint) == 42:
                                            self.toPrint.append(uid7['userid'])
                                            self.toPrint.append(uid7['name'])
                                            self.toPrint.append(uid7['worksfor'])
                                            self.toPrint.append(uid7['role'])
                                            self.toPrint.append(uid7['location'])
                                            self.toPrint.append(uid7['orgName'])
                                            self.writeXlsx(self.line,self.toPrint)            
                                        elif len(self.toPrint) > 42:
                                            self.toPrint = self.toPrint[:42]
                                            self.toPrint.append(uid7['userid'])
                                            self.toPrint.append(uid7['name'])
                                            self.toPrint.append(uid7['worksfor'])
                                            self.toPrint.append(uid7['role'])
                                            self.toPrint.append(uid7['location'])
                                            self.toPrint.append(uid7['orgName'])
                                            self.writeXlsx(self.line,self.toPrint)
                                        Level8 = self.getSubs(uid7['userid'])
                                        for uid8 in Level8:
                                            if self.maxlevel<=8:
                                                self.maxlevel=8
                                            if len(self.toPrint) == 48:
                                                self.toPrint.append(uid8['userid'])
                                                self.toPrint.append(uid8['name'])
                                                self.toPrint.append(uid8['worksfor'])
                                                self.toPrint.append(uid8['role'])
                                                self.toPrint.append(uid8['location'])
                                                self.toPrint.append(uid8['orgName'])
                                                self.writeXlsx(self.line,self.toPrint)            
                                            elif len(self.toPrint) > 48:
                                                self.toPrint = self.toPrint[:48]
                                                self.toPrint.append(uid8['userid'])
                                                self.toPrint.append(uid8['name'])
                                                self.toPrint.append(uid8['worksfor'])
                                                self.toPrint.append(uid8['role'])
                                                self.toPrint.append(uid8['location'])
                                                self.toPrint.append(uid8['orgName'])
                                                self.writeXlsx(self.line,self.toPrint)
                                            Level9 = self.getSubs(uid8['userid'])
                                            for uid9 in Level9:
                                                if self.maxlevel<=9:
                                                    self.maxlevel=9
                                                if len(self.toPrint) == 54:
                                                    self.toPrint.append(uid9['userid'])
                                                    self.toPrint.append(uid9['name'])
                                                    self.toPrint.append(uid9['worksfor'])
                                                    self.toPrint.append(uid9['role'])
                                                    self.toPrint.append(uid9['location'])
                                                    self.toPrint.append(uid9['orgName'])
                                                    self.writeXlsx(self.line,self.toPrint)            
                                                elif len(self.toPrint) > 54:
                                                    self.toPrint = self.toPrint[:54]
                                                    self.toPrint.append(uid9['userid'])
                                                    self.toPrint.append(uid9['name'])
                                                    self.toPrint.append(uid9['worksfor'])
                                                    self.toPrint.append(uid9['role'])
                                                    self.toPrint.append(uid9['location'])
                                                    self.toPrint.append(uid9['orgName'])
                                                    self.writeXlsx(self.line,self.toPrint)
        self.outWB.save(self.outputFilename)
        print(f"{datetime.datetime.now()} Saving output file in {self.outputFilename}")
        return None

def main():
    source = sys.argv[1]
    output = sys.argv[2]
    if "xlsx" in source and "xlsx" in output:
        newExec = OrgTable(source,output)
    else:
        print("!INFO: No valid inputs provided using the values set by default")
        newExec = OrgTable("./Heinz GALTest.xlsx","./output.xlsx")
    newExec.ProcessData()
    newExec.addHeaders()

if __name__ == "__main__":
    main()